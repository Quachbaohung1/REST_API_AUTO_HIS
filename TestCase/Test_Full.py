from unittest import TestCase
from TestCase import Test_Tiếp_nhận, Test_Khám_bệnh, Test_Kê_toa, Test_Viện_phí, Test_Dược
import unittest
from openpyxl import load_workbook
import logging


# # Xác định đường dẫn đến file log
# log_file_path = 'D:/HIS api automation/Result/TestResult_Full.txt'
#
# # Xóa file log cũ nếu tồn tại
# with open(log_file_path, 'w'):  # Mở file trong chế độ ghi để xóa nội dung
#     pass  # Pass đơn giản làm gì để xóa
#
# # Thiết lập logging để ghi vào file mới
# logging.basicConfig(filename=log_file_path, level=logging.INFO)
#
#
# # Ghi lại kết quả của test
# def log_test_result(test_name, result):
#     logging.info(f'Test "{test_name}" - Result: {result}')


def read_test_cases():
    test_case_file = "D://testScenario.xlsx"
    try:
        wb = load_workbook(filename=test_case_file, read_only=True)
        sheet = wb.active
    except FileNotFoundError:
        print(f"File '{test_case_file}' không tồn tại.")
        return []

    test_cases = []
    current_module = None
    test_case_description = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        module_name = row[1] if row[1] else current_module
        test_case_description = row[3] if row[3] else test_case_description
        priority = row[4]
        other_info = row[6]

        if module_name:
            current_module = module_name

        test_case = {
            'Module': current_module,
            'Test Scenario Description': test_case_description,
            'Test Case Id': row[5],
            'Priority': priority,
            'Test Case Description': other_info
        }
        test_cases.append(test_case)

    print(test_cases)
    return test_cases


# Ánh xạ testcaseid với tên hàm test tương ứng
test_case_mapping_TN = {
    'TC_01': 'test_case_01',
    'TC_02': 'test_case_02',
    'TC_03': 'test_case_03',
    'TC_04': 'test_case_04',
    'TC_05': 'test_case_05',
    'TC_06': 'test_case_06',
    'TC_07': 'test_case_07',
    'TC_08': 'test_case_08',
    'TC_09': 'test_case_09',
    'TC_10': 'test_case_10',
    'TC_11': 'test_case_11',
    'TC_12': 'test_case_12',
    'TC_13': 'test_case_01',
    'TC_14': 'test_case_03',
    'TC_15': 'test_case_04',
    'TC_16': 'test_case_05',
    'TC_17': 'test_case_06',
    'TC_18': 'test_case_07',
    'TC_19': 'test_case_09',
    'TC_20': 'test_case_10',
    'TC_21': 'test_case_11',
    'TC_22': 'test_case_13',
    'TC_23': 'test_case_13',
    'TC_24': 'test_case_14'
}

test_case_mapping_CDDV = {
    'TC_01': 'test_case_01',
    'TC_02': 'test_case_02',
    'TC_03': 'test_case_03',
    'TC_04': 'test_case_04',
    'TC_05': 'test_case_05'
}


test_case_mapping_Medicine = {
    'TC_01': 'test_case_01',
    'TC_02': 'test_case_02',
    'TC_03': 'test_case_03',
    'TC_04': 'test_case_04',
    'TC_05': 'test_case_05',
    'TC_06': 'test_case_06',
    'TC_07': 'test_case_07',
    'TC_08': 'test_case_08',
    'TC_09': 'test_case_09',
    'TC_10': 'test_case_10',
    'TC_11': 'test_case_11',
    'TC_12': 'test_case_12',
    'TC_13': 'test_case_13',
    'TC_14': 'test_case_14',
    'TC_15': 'test_case_02',
    'TC_16': 'test_case_03',
    'TC_17': 'test_case_04',
    'TC_18': 'test_case_05',
    'TC_19': 'test_case_06',
    'TC_20': 'test_case_07',
    'TC_21': 'test_case_08',
    'TC_22': 'test_case_09',
    'TC_23': 'test_case_10',
    'TC_24': 'test_case_11',
    'TC_25': 'test_case_12',
    'TC_26': 'test_case_14'
}


test_case_mapping_VP = {
    'TC_01': 'test_case_01',
    'TC_02': 'test_case_02',
    'TC_03': 'test_case_03'
}


test_case_mapping_Store = {
    'TC_01': 'test_case_01',
    'TC_02': 'test_case_02'
}


# class LogTestResult(unittest.TextTestResult):
#     def addSuccess(self, test):
#         super().addSuccess(test)
#         log_test_result(test.id(), 'PASSED')
#
#     def addFailure(self, test, err):
#         super().addFailure(test, err)
#         log_test_result(test.id(), 'FAILED')
#
#     def addError(self, test, err):
#         super().addError(test, err)
#         log_test_result(test.id(), 'ERROR')
#
#     def addSkip(self, test, reason):
#         super().addSkip(test, reason)
#         log_test_result(test.id(), f'SKIPPED ({reason})')


def run_tests(test_cases):
    suite = unittest.TestSuite()

    for test_case in test_cases:
        module = test_case['Module']
        testcaseid = test_case['Test Case Id']

        # Thêm các test case vào suite dựa trên module và priority
        if module == 'Tiếp nhận':
            if testcaseid in test_case_mapping_TN:
                test_function_name = test_case_mapping_TN[testcaseid]
                suite.addTest(Test_Tiếp_nhận.TestProcessPatient(test_function_name))
            else:
                print(f"Không tìm thấy hàm test cho testcaseid '{testcaseid}'.")
        elif module == 'Khám bệnh - CDDV':
            if testcaseid in test_case_mapping_CDDV:
                test_function_name = test_case_mapping_CDDV[testcaseid]
                suite.addTest(Test_Khám_bệnh.TestProcessExamination(test_function_name))
            else:
                print(f"Không tìm thấy hàm test cho testcaseid '{testcaseid}'.")
        elif module == 'Khám bệnh - Kê toa':
            if testcaseid in test_case_mapping_Medicine:
                test_function_name = test_case_mapping_Medicine[testcaseid]
                suite.addTest(Test_Kê_toa.TestProcessPrescription(test_function_name))
            else:
                print(f"Không tim thấy hàm test cho testcaseid '{testcaseid}'.")
        elif module == 'Viện phí':
            if testcaseid in test_case_mapping_VP:
                test_function_name = test_case_mapping_VP[testcaseid]
                suite.addTest(Test_Viện_phí.TestProcessCosts(test_function_name))
            else:
                print(f"Không tìm thấy hàm test cho testcaseid '{testcaseid}'.")
        elif module == 'Dược':
            if testcaseid in test_case_mapping_Store:
                test_function_name = test_case_mapping_Store[testcaseid]
                suite.addTest(Test_Dược.TestProcessStore(test_function_name))
            else:
                print(f"Không tìm thấy hàm test cho testcaseid '{testcaseid}'.")

    # Thực thi suite
    # resultclass = LogTestResult
    runner = unittest.TextTestRunner()
    result = runner.run(suite)

    # Kiểm tra kết quả và dừng nếu có lỗi với priority là 'P'
    if any(test_case['Priority'] == 'P' and not result.wasSuccessful() for test_case in test_cases):
        print("Stopped further tests due to failure in high-priority test cases.")
        return False
    return True


if __name__ == '__main__':
    test_cases = read_test_cases()
    success = run_tests(test_cases)
